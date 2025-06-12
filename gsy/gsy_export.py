"""
gsy_export.py
Data export and processing module for GSY data extraction system.

This module handles all data export operations for both DTR and GTP systems including:
- DataFrame column reordering to standard format
- Data validation and error marking
- Excel file generation with formatting and macros
- Template-based Excel export with VBA preservation
- User interaction for file overwrite confirmation

The system type (DTR or GTP) is determined by the SCRIPT_SYSTEM constant
which should be set in the calling script (getDTRs.py or getGTPs.py).

Dependencies:
- pandas: For DataFrame operations
- openpyxl: For Excel file manipulation
- shutil: For file operations

"""

import os
import sys
import time
import shutil
import logging
from typing import Optional
from pandas import DataFrame
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment

# System type constant - should be set by calling script
# Valid values: 'DTR' or 'GTP'
try:
    from __main__ import SCRIPT_SYSTEM
except ImportError:
    # Default to None if not set - will raise error if functions are called
    SCRIPT_SYSTEM = None

# System enviroment constant - should be set by calling script
# Valid values: 'QAS' or 'Production'
try:
    from __main__ import SCRIPT_ENV
except ImportError:
    # Default to None if not set - will raise error if functions are called
    SCRIPT_ENV= None

# Standard column order for DTR system
STANDARD_COLUMN_ORDER_DTR = [
    'program',
    'gtpNumber',
    'gtpRevision',
    'gtpTitle',
    'gtpStatus',
    'technology',
    'gtpIssuedDate',
    'gtpApprovalDate',
    'gtpSubmittalDate',
    'gtpFinishedDate',
    'dtrRequiredDate',
    'gtpApprovedAgreedDate',
    'gtpAgreedDateInDiscussion',
    'testVehicle',
    'category',
    'pep',
    'MHPlannedTest',
    'testOwner',
    'testResponsible',
    'supervisor',
    'docType',
    'docNumber',
    'docRevision',
    'docTitle',
    'docStatus',
    'docSubmittalDate',
    'docSubmitalToApprovalDate',
    'docApprovalDate',
    'docAuthor',
    'docNextApprover',
    'docApprovers',
    'MH',
    'duracao',
    'dtrStatus',
    'errorMsg'
]

# Standard column order for GTP system
STANDARD_COLUMN_ORDER_GTP = [
    'program',
    'gtpNumber',
    'gtpRevision',
    'gtpTitle',
    'technology',
    'gtpStatus',
    'issueDate',
    'approvalDate',
    'submittalDate',
    'finishedDate',
    'agreedDate',
    'supervisor',
    'testVehicle',
    'responsible',
    'dtrStatus',
    'dtrCount',
    'dtrQuantity',
    'errorMsg'
]

# Excel formatting constants
HEADER_FILL_COLOR = 'D3D3D3'
DATE_FORMAT = 'dd/mm/yyyy'
MAX_COLUMN_WIDTH = 50
MIN_SAMPLE_ROWS = 100  # Number of rows to sample for column width calculation

def _validate_system():
    """Validate that SCRIPT_SYSTEM is properly set."""
    if SCRIPT_SYSTEM not in ['DTR', 'GTP']:
        raise ValueError(
            f"SCRIPT_SYSTEM must be set to 'DTR' or 'GTP' in calling script. "
            f"Current value: {SCRIPT_SYSTEM}"
        )

def _get_standard_columns():
    """Get the appropriate standard column order based on system type."""
    _validate_system()
    if SCRIPT_SYSTEM == 'DTR':
        return STANDARD_COLUMN_ORDER_DTR
    else:  # GTP
        return STANDARD_COLUMN_ORDER_GTP

def checkExcel(filename: str = "BDI_TechReports.xlsm") -> None:
    """
    Check if Excel file exists and handle deletion with user confirmation.
    
    This function provides interactive user prompts to handle existing files,
    including retry logic for files that are currently open in another application.
    
    Args:
        filename: Name of the Excel file to check (default: BDI_TechReports.xlsm)
    
    Returns:
        None
        
    Raises:
        SystemExit: If user cancels operation or unrecoverable error occurs
    """
    # Skip if file doesn't exist
    if not os.path.exists(filename):
        logging.info(f"File {filename} does not exist. Proceeding with export...")
        return
    
    # Request user confirmation for deletion
    print(f"\nFile '{filename}' already exists.")
    
    while True:
        response = input("Do you want to delete it? (yes/no): ").strip().lower()
        
        if response in ['yes', 'y']:
            break
        elif response in ['no', 'n']:
            print("Operation cancelled by user. Exiting program...")
            sys.exit(0)
        else:
            print("Please answer 'yes' or 'no'.")
    
    # Attempt file deletion with retry logic
    delete_attempts = 0
    while os.path.exists(filename):
        try:
            os.remove(filename)
            print(f"File '{filename}' deleted successfully.")
            logging.info(f"File {filename} deleted successfully")
            return
            
        except PermissionError:
            delete_attempts += 1
            if delete_attempts == 1:
                print(f"\nCannot delete '{filename}' - file may be in use.")
                print("Please close the file if it's open in Excel or another application.")
            
            # Provide user options
            print("\nOptions:")
            print("1. Press Enter to try again")
            print("2. Type 'exit' to quit the program")
            
            user_input = input("Your choice: ").strip().lower()
            
            if user_input == 'exit':
                print("Exiting program...")
                sys.exit(0)
            
            # Brief pause before retry
            time.sleep(0.5)
            
        except Exception as e:
            logging.error(f"Unexpected error deleting {filename}: {e}")
            print(f"Unexpected error occurred: {e}")
            print("Exiting program...")
            sys.exit(1)

def dfReorder(df: DataFrame) -> DataFrame:
    """
    Reorder DataFrame columns according to standard column order.
    
    This function ensures consistent column ordering for export, adding
    missing columns as needed and preserving any additional columns
    not in the standard order. The standard order used depends on the
    SCRIPT_SYSTEM setting (DTR or GTP).
    
    Args:
        df: Input DataFrame to reorder
        
    Returns:
        DataFrame: New DataFrame with reordered columns
        
    Note:
        - Missing standard columns are added with None values
        - Extra columns are appended after standard columns
        - Original DataFrame is not modified
    """
    # Get appropriate standard column order
    standard_column_order = _get_standard_columns()
    
    # Create a copy to avoid modifying original
    df_copy = df.copy()
    
    # Add missing standard columns
    for col in standard_column_order:
        if col not in df_copy.columns:
            df_copy[col] = None
    
    # Identify columns in both standard order and DataFrame
    existing_standard_columns = [
        col for col in standard_column_order 
        if col in df_copy.columns
    ]
    
    # Identify additional columns not in standard order
    additional_columns = [
        col for col in df_copy.columns 
        if col not in standard_column_order
    ]
    
    # Combine columns: standard first, then additional
    final_column_order = existing_standard_columns + additional_columns
    
    return df_copy[final_column_order]

def dfCheck(df: DataFrame) -> None:
    """
    Validate DataFrame and mark rows with errors according to business rules.
    
    Rules implemented:
    - DTR: dtrStatus is 'Submitted' or 'Approved' but gtpSubmittalDate is empty
    - GTP: dtrStatus is 'Submitted' or 'Approved' but submittalDate is empty
    
    Args:
        df: DataFrame to validate (modified in-place)
        
    Returns:
        None
        
    Note:
        Results are stored in the 'errorMsg' column
    """
    _validate_system()
    
    # Initialize errorMsg column if not present
    if 'errorMsg' not in df.columns:
        df['errorMsg'] = ''
    
    # Determine which date column to check based on system
    if SCRIPT_SYSTEM == 'DTR':
        date_column = 'gtpSubmittalDate'
        error_message = "gtpSubmittalDate empty and dtrStatus in (Submitted, Approved)"
    else:  # GTP
        date_column = 'submittalDate'
        error_message = "submittalDate empty and dtrStatus in (Submitted, Approved)"
    
    # Apply validation rule
    if 'dtrStatus' in df.columns and date_column in df.columns:
        mask = (
            df['dtrStatus'].isin(['Submitted', 'Approved']) & 
            (df[date_column].isna() | (df[date_column] == ''))
        )
        df.loc[mask, 'errorMsg'] = error_message
        
        # Log validation results
        error_count = mask.sum()
        if error_count > 0:
            logging.info(
                f"Data validation ({SCRIPT_SYSTEM}): Found {error_count} rows with "
                f"missing {date_column}"
            )

def dfExportExcel(df: DataFrame, filename: str, api_key: Optional[str] = None, 
                  template_path: Optional[str] = './gsy/template.xlsm') -> None:
    """
    Export DataFrame to Excel file using macro-enabled template with formatting.
    
    This function creates a formatted Excel file by:
    1. Copying a macro-enabled template
    2. Writing data with proper formatting
    3. Storing API key in hidden configuration sheet
    4. Applying styles, filters, and column sizing
    
    Args:
        df: DataFrame to export
        filename: Output filename (will be converted to .xlsm if needed)
        api_key: Optional API key to store in hidden Cfg sheet
        template_path: Path to Excel template file (uses system default if None)
        
    Returns:
        None
        
    Raises:
        FileNotFoundError: If template file is not found
        Exception: For other Excel operation errors
        
    Note:
        Requires openpyxl with macro support
    """

    if not os.path.isfile(template_path):
        logging.error(f"Error: {template_path} not found")
        logging.error("This file is required for create an excel file with macros")
        sys.exit(1)
    
    # Ensure macro-enabled extension
    if not filename.endswith('.xlsm'):
        logging.error(f"Error: {template_path} is not in the .xlsm format")
        logging.error("This file is required for create an excel file with macros")
        sys.exit(1)
    
    logging.info(f"Exporting to Excel file: {filename} (System: {SCRIPT_SYSTEM})")
    
    # Copy template to destination
    try:
        shutil.copy2(template_path, filename)
        logging.info(f"Template copied from {template_path}")
    except FileNotFoundError:
        logging.error(f"Template file not found: {template_path}")
        raise
    except Exception as e:
        logging.error(f"Error copying template: {e}")
        raise
    
    # Load workbook preserving VBA code
    try:
        workbook = load_workbook(filename, keep_vba=True)
    except Exception as e:
        logging.error(f"Error loading workbook: {e}")
        raise
    
    # Store API key, script_system and script_env in hidden Cfg 
    if 'Cfg' in workbook.sheetnames:
        cfg_sheet = workbook['Cfg']
        cfg_sheet.cell(row=1, column=2, value=api_key)  # Cell B1
        cfg_sheet.cell(row=2, column=2, value=SCRIPT_SYSTEM)  # Cell B2
        cfg_sheet.cell(row=3, column=2, value=SCRIPT_ENV)  # Cell B3
        
        cfg_sheet.sheet_state = 'hidden'
        logging.info("API key and Script SYSTEM stored in hidden Cfg sheet")
    else:
        logging.warning("Cfg sheet not found in template")
    
    # Prepare data worksheet
    worksheet = _prepare_data_sheet(workbook)
    
    # Write data to worksheet
    _write_data_to_sheet(worksheet, df)
    
    # Apply formatting
    _apply_sheet_formatting(worksheet, df)
    
    # Save workbook
    try:
        workbook.save(filename)
        workbook.close()
        logging.info(f"Excel file saved successfully: {filename}")
        logging.info(f"Total rows exported: {len(df):,}")
    except Exception as e:
        logging.error(f"Error saving workbook: {e}")
        raise

def _prepare_data_sheet(workbook):
    """Prepare or create the Data worksheet."""
    if 'Data' in workbook.sheetnames:
        worksheet = workbook['Data']
        # Clear existing content
        worksheet.delete_rows(1, worksheet.max_row)
    else:
        worksheet = workbook.create_sheet('Data')
    
    return worksheet

def _write_data_to_sheet(worksheet, df: DataFrame) -> None:
    """Write DataFrame data to worksheet."""
    # Write headers
    for col_num, column_title in enumerate(df.columns, 1):
        cell = worksheet.cell(row=1, column=col_num)
        cell.value = column_title
    
    # Write data rows
    for row_num, row_data in enumerate(df.itertuples(index=False), 2):
        for col_num, value in enumerate(row_data, 1):
            cell = worksheet.cell(row=row_num, column=col_num)
            cell.value = value

def _apply_sheet_formatting(worksheet, df: DataFrame) -> None:
    """Apply formatting to the worksheet."""
    # Freeze header row
    worksheet.freeze_panes = 'A2'
    
    # Create header styles
    header_font = Font(bold=True)
    header_fill = PatternFill(
        start_color=HEADER_FILL_COLOR, 
        end_color=HEADER_FILL_COLOR, 
        fill_type='solid'
    )
    
    # Apply header formatting
    for cell in worksheet[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # Add AutoFilter
    worksheet.auto_filter.ref = worksheet.dimensions
    
    # Format date columns
    _format_date_columns(worksheet, df)
    
    # Auto-size columns
    _autosize_columns(worksheet, df)

def _format_date_columns(worksheet, df: DataFrame) -> None:
    """Apply date formatting to columns containing dates."""
    date_columns = [
        i for i, col_name in enumerate(df.columns) 
        if 'date' in col_name.lower()
    ]
    
    for col_idx in date_columns:
        col_letter = worksheet.cell(row=1, column=col_idx + 1).column_letter
        for row in range(2, len(df) + 2):
            cell = worksheet[f'{col_letter}{row}']
            cell.number_format = DATE_FORMAT

def _autosize_columns(worksheet, df: DataFrame) -> None:
    """Automatically adjust column widths based on content."""
    for i, column in enumerate(df.columns):
        # Sample cells for width calculation
        sample_size = min(MIN_SAMPLE_ROWS, len(df) + 1)
        column_cells = [
            worksheet.cell(row=r, column=i + 1) 
            for r in range(1, sample_size + 1)
        ]
        
        # Calculate maximum content length
        max_length = max(
            len(str(cell.value or '')) 
            for cell in column_cells
        ) + 2  # Add padding
        
        # Apply width with maximum limit
        col_letter = worksheet.cell(row=1, column=i + 1).column_letter
        worksheet.column_dimensions[col_letter].width = min(max_length, MAX_COLUMN_WIDTH)